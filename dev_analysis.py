#!/usr/bin/env python3

import argparse
import glob
import networkx as nx
import os
import pickle
import sys
from graphs import call_graph
from graphs.devrank import devrank
from graphs.pagerank import pagerank
from graphs.call_commit_graph import CallCommitGraph
from lxml import etree
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# Utility functions

def add_fname(item, func2file):
    fname = func2file[item[0]] if item[0] in func2file else 'Unknown'
    return [item[0], item[1], fname]

def fname_array(ranks_array, func2file):
    array = []
    for ranks in ranks_array:
        ranks = sorted(ranks, key=lambda r: r[1], reverse=True)
        array.append([add_fname(r, func2file) for r in ranks])
    return array

# Prepare the call graph

def build_call_graph(xml_dir):
    print('Building the call graph for ' + xml_dir)

    graph = nx.DiGraph()
    func2file = {}

    for xml in glob.glob(xml_dir + '/**/*.[ch].xml', recursive=True):
        tree = etree.parse(xml)
        roots = [tree.getroot()]
        _, f2f = call_graph.c.build_call_graph(roots, {}, G=graph)
        for func, fname in f2f.items():
            func2file[func] = fname
    print("Number of nodes: {}".format(len(graph.nodes())))
    print("Number of edges: {}".format(len(graph.edges())))
    n = nx.number_weakly_connected_components(graph)
    print("Number of connected components: {}".format(n))
    return graph, func2file

def pagerank_c(graph, alpha_low, alpha_high, alpha_step):
    pra = []
    a = alpha_low
    while (a <= alpha_high):
        print('Running PageRank-C for alpha=' + str(a))
        pr = pagerank(graph, alpha=a)
        pra.append(pr.items())
        a += alpha_step
    return pra

def devrank_c(graph, alpha_low, alpha_high, alpha_step):
    dra = []
    a = alpha_low
    while (a <= alpha_high):
        print('Running DevRank-C for alpha=' + str(a))
        dr = devrank(graph, alpha=a)
        dra.append(dr.items())
        a += alpha_step
    return dra

def devrank_c2(graph, alpha_low, alpha_high, alpha_step):
    dra = []
    a = alpha_low
    while (a <= alpha_high):
        print('Running DevRank-CC for alpha=' + str(a))
        dra.append(graph.devrank_functions(a))
        a += alpha_step
    return dra 

def output(ranks_array, worksheet, alpha_low, alpha_high, alpha_step):
    a = alpha_low
    c = 1
    for ranks in ranks_array:
        worksheet.cell(column=c, row=1, value='alpha='+ str(a))
        for r, item in enumerate(ranks):
            for i, v in enumerate(item):
                worksheet.cell(column=c+i, row=r+2, value=item[i])
        a += alpha_step
        c += 3

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-x', required=True,
                        help='dir of the XML form of source code')
    parser.add_argument('-s', help='dir of the source code repo') 
    parser.add_argument('-o', required=True, help='output .xmlx file')
    parser.add_argument('-a', type=float, nargs=3, required=True,
                        help='lowest alpha, highest alpha and \
                              step for iteration')
    parser.add_argument('-n', type=int, nargs='+',
                        help='(multiple) # of commits to analyze') 
    parser.add_argument('-p', action='store_true',
                        help='run PageRank on the call graph')
    parser.add_argument('-d', action='store_true', 
                        help='run DevRank on the call graph')
    parser.add_argument('-c', action='store_true',
                        help='run DevRank on the call-commit graph')
    args = parser.parse_args()

    xml_dir = os.path.expanduser(args.x)
    if not os.path.isdir(xml_dir):
        sys.exit('Invalid dir for -x.')
 
    wb = load_workbook(args.o) if os.path.isfile(args.o) else Workbook() 
    graph = None
    func2file = None

    if args.p:
        if graph is None:
            graph, func2file = build_call_graph(xml_dir)
        ranks_array = pagerank_c(graph, args.a[0], args.a[1], args.a[2])
        ranks_array = fname_array(ranks_array, func2file)
        ws = wb.create_sheet('PageRank-C')
        output(ranks_array, ws, args.a[0], args.a[1], args.a[2])
    if args.d:
        if graph is None:
            graph, func2file = build_call_graph(xml_dir)
        ranks_array = devrank_c(graph, args.a[0], args.a[1], args.a[2])
        ranks_array = fname_array(ranks_array, func2file)
        ws = wb.create_sheet('DevRank-C')
        output(ranks_array, ws, args.a[0], args.a[1], args.a[2])
    if args.c:
        if args.s is None or args.n is None:
            sys.exit('Specify -s and -n for call-commit graph.')
        src_dir = os.path.expanduser(args.s)
        if not os.path.isdir(src_dir):
            sys.exit('Invalid dir for -s.')
        if graph is None:
            graph, func2file = build_call_graph(xml_dir) 
        ccg = CallCommitGraph(src_dir)
        for n in args.n:
            ccg.process(rev='v4.10', num_commits=n)
            ranks_array = devrank_c2(ccg, args.a[0], args.a[1], args.a[2])
            ranks_array = fname_array(ranks_array, func2file)
            ws = wb.create_sheet('DevRank-CC n=' + str(n))
            output(ranks_array, ws, args.a[0], args.a[1], args.a[2])

    if not args.p and not args.d and not args.c:
        sys.exit('Specify -p, -d or -c.')
    else:
        wb.save(args.o)

if __name__ == '__main__':
    main()
