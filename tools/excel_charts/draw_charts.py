#!/usr/bin/env python3

import argparse
import distance
import excel
from openpyxl.utils import get_column_letter
from gini.gini import gini
from openpyxl import load_workbook
from openpyxl.chart import LineChart
import math
import numpy
from openpyxl.chart import Reference
from openpyxl.chart import Series
from openpyxl import Workbook

def alpha(worksheet, i):
    a = excel.cell(worksheet, (1, 1), step_column=3, index=i).value
    if not a: return None
    assert a.startswith('alpha=')
    return a[len('alpha='):][0:4]

def draw_ordered_ranks(worksheet):
    chart = LineChart()
    chart.title = 'Ranks in descending order'
    chart.x_axis.title = 'Function'
    chart.y_axis.title = 'Rank'
    chart.style = 11

    i = 0
    while True:
        a = alpha(worksheet, i)
        if a is None: break
        col = 3 * i + 2
        data = Reference(worksheet, min_col=col, max_col=col,
                         min_row=2, max_row=min(worksheet.max_row, 100))
        series = Series(data, title='a=' + a)
        series.smooth = True
        series.graphicalProperties.line.width = 20000
        chart.append(series)
        i += 1
    worksheet.add_chart(chart, 'D25')

def calc_stats(worksheet):
    alphas = [['alpha']]
    stats = [[worksheet.title + ' rank distance',
              worksheet.title + ' value distance',
              worksheet.title + ' pair changes',
              worksheet.title + ' gini']]

    map1 = None
    i = 0
    while True:
        a = alpha(worksheet, i)
        if a is None: break    
        map2 = {}
        values = []
        for cell in worksheet[get_column_letter(3 * i + 1)][1:]:
            rank = cell.row - 1
            #TODO: epsilon=1e-5
            value = worksheet.cell(row=cell.row, column=cell.col_idx + 1).value
            map2[cell.value] = [rank, float(value)]
            values.append(float(value))
        if map1 is None:
            map1 = map2
        else:
            alphas.append([a])
            rank_d = distance.deviation(map1, map2, 0)
            value_d = distance.deviation(map1, map2, 1)
            pc = distance.pair_changes(map1, map2, 0)
            gc = gini(numpy.array(values))
            stats.append([rank_d, value_d, pc, gc])
            map1 = map2
        i += 1
    return alphas, stats 

def draw_stats(workbook):
    stats_sheet = excel.sheet(workbook, 'Stats')
    if not stats_sheet:
        stats_sheet = workbook.create_sheet('Stats')

    data_col = 2 # the starting distance data column
    for ws in workbook.worksheets:
        if ws.title == 'Stats':
            continue

        alphas, stats = calc_stats(ws)

        # Set or check the alpha column.
        if data_col == 2:
            excel.fillout(stats_sheet, (1, 1), alphas)
        else:
            a = excel.fillin(stats_sheet, (1, 1), len(alphas), 1)
            assert a == alphas

        # Set the distance column(s) for the current worksheet.
        excel.fillout(stats_sheet, (1, data_col), stats)

        data_col += len(stats[0])

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help='Excel file to draw charts in')
    args = parser.parse_args()

    wb = load_workbook(args.file)

    for ws in wb.worksheets:
        if ws.title == 'Stats': continue
        print('Drawing for ' + ws.title)
        draw_ordered_ranks(ws)

    draw_stats(wb)

    wb.save(args.file)

if __name__ == '__main__':
    main()

