#!/usr/bin/env python3

from openpyxl import Workbook

def fillout(sheet, position, data):
    row = position[0]
    column = position[1]
    for i, array in enumerate(data):
        for j, value in enumerate(array):
            sheet.cell(row=row+i, column=column+j, value=value)

def fillin(sheet, position, num_rows, num_columns):
    row = position[0]
    column = position[1] 
    data = [[None for j in range(num_columns)] for i in range(num_rows)]
    for i in range(num_rows):
        for j in range(num_columns):
            data[i][j] = sheet.cell(row=row+i, column=column+j).value
    return data

def cell(sheet, position, step_row=0, step_column=0, index=0):
    row = position[0]
    column = position[1]
    return sheet.cell(row=row + index * step_row,
                      column=column + index * step_column)

def sheet(workbook, sheetname):
    try:
        return workbook[sheetname]
    except KeyError:
        return None

def main():
    wb = Workbook()
    ws = wb.active
    data = [[x] for x in range(10)]
    fillout(ws, (1, 1), data)

    data = [[x, 2 * x] for x in range(10)]
    fillout(ws, (2, 2), data)

    data = [['Sheet1 rank distance', 'Sheet1 value distance'], [0.0, 0.5], [1.0, 0.5]]
    fillout(ws, (1,2), data)

    print(sheet(wb, 'Sheet'))
    print(sheet(wb, 'InvalidSheetName'))

    print(fillin(ws, (1, 1), 10, 1))
    print(fillin(ws, (2, 2), 10, 2))

    i = 0
    while True:
        c = cell(ws, (2, 1), step_column=2, index=i)
        if c.value is None: break
        print(c.value)
        i += 1

    wb.save('check.xlsx')

if __name__ == '__main__':
    main()

