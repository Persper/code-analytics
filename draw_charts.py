#!/usr/bin/env python3

import argparse
# import editdistance
import math
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import LineChart
from openpyxl.chart import Reference
from openpyxl.chart import Series
from openpyxl.utils import get_column_letter

def alpha(worksheet, i):
    a = worksheet.cell(row=1, column=3 * i + 1).value
    if a is None or not a.startswith('alpha='):
        return None
    return a[len('alpha='):][0:4]

def draw_ordered_ranks(worksheet):
    chart = LineChart()
    chart.title = 'Ranks in descending order'
    chart.x_axis.title = 'Function'
    chart.y_axis.title = 'Rank'
    chart.style = 11

    i = 1
    while True:
        a = alpha(worksheet, i)
        if a is None: break
        col = 3 * i + 2
        data = Reference(worksheet, min_col=col, max_col=col,
                         min_row=2, max_row=min(worksheet.max_row, 100))
        series = Series(data, title='a=' + a)
        series.smooth = True
        series.graphicalProperties.line.width = 20000
        chart.append(series)
        i += 2
    worksheet.add_chart(chart, 'D25')

def distance(map1, map2):
    d = 0
    for key, value in map1.items():
        d += (value - map2.get(key, value))**2
    return d

def draw_distance(worksheet):
    alphas = []
    maps = []
    # seqs = []
    i = 0
    while True:
        a = alpha(worksheet, i)
        if a is None: break
        m = dict()
        s = [] 
        for cell in worksheet[get_column_letter(3 * i + 1)]:
            m[cell.value] = cell.row
            # s.append(cell.value)
        alphas.append(a)
        maps.append(m)
        # seqs.append(s)
        i += 1
    col_alpha = 3 * i + 1
    col_distance = 3 * i + 2
    worksheet.cell(column=col_alpha, row=1, value='alpha')
    worksheet.cell(column=col_distance, row=1, value='distance')
    r = 2
    while r < i:
        worksheet.cell(column=col_alpha, row=r, value=alphas[r - 1])
        d = distance(maps[r - 1], maps[r])
        # d = editdistance.eval(seqs[r - 1], seqs[r])
        d = math.sqrt(d / worksheet.max_row)
        worksheet.cell(column=col_distance, row=r, value=d)
        r += 1

def draw_unknown(worksheet):
    return

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help='Excel file to draw charts in')
    args = parser.parse_args()

    wb = load_workbook(args.file)

    for ws in wb.worksheets:
        print('Drawing for ' + ws.title)
        draw_ordered_ranks(ws)
        draw_distance(ws)

    wb.save(args.file)

if __name__ == '__main__':
    main()
